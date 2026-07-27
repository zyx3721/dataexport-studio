from itertools import chain, islice
from pathlib import Path
from typing import Callable, Iterable, Optional, Sequence

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from ...domain.errors import ExportError


EXCEL_DATA_ROWS_PER_SHEET = 1_048_575
COLUMN_WIDTH_SAMPLE_SIZE = 200
HEADER_FONT = Font(name="SimSun", size=11, bold=True)
DATA_FONT = Font(name="SimSun", size=11)
CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(*(Side(style="thin") for _ in range(4)))
HEADER_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="medium"),
    bottom=Side(style="thin"),
)


def write_workbook(
    destination: Path,
    headers: Sequence[str],
    rows: Iterable[tuple],
    max_rows: int,
    on_progress: Optional[Callable[[int], None]] = None,
    is_cancelled: Optional[Callable[[], bool]] = None,
) -> int:
    if max_rows < 1:
        raise ExportError("最大导出行数必须大于零。")
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_suffix(".partial.xlsx")
    workbook = Workbook(write_only=True)
    count = 0
    sheet_index = 1
    try:
        sample_rows = list(islice(rows, COLUMN_WIDTH_SAMPLE_SIZE))
        column_widths = _column_widths(headers, sample_rows)
        sheet = _create_sheet(workbook, sheet_index, headers, column_widths)
        for row in chain(sample_rows, rows):
            if is_cancelled and is_cancelled():
                raise ExportError("导出已取消。")
            if count >= max_rows:
                break
            if count and count % EXCEL_DATA_ROWS_PER_SHEET == 0:
                sheet_index += 1
                sheet = _create_sheet(workbook, sheet_index, headers, column_widths)
            sheet.append([_data_cell(sheet, value) for value in row])
            count += 1
            if on_progress and count % 100 == 0:
                on_progress(count)
        workbook.save(temporary)
        temporary.replace(destination)
        if on_progress:
            on_progress(count)
        return count
    except ExportError:
        raise
    except (OSError, TypeError, ValueError) as exc:
        raise ExportError("无法写入 Excel 文件。") from exc
    finally:
        workbook.close()
        if temporary.exists():
            temporary.unlink()


def _create_sheet(workbook: Workbook, index: int, headers: Sequence[str], column_widths: Sequence[int]):
    sheet = workbook.create_sheet("数据_{}".format(index))
    sheet.sheet_view.showGridLines = True
    for number, width in enumerate(column_widths, 1):
        sheet.column_dimensions[get_column_letter(number)].width = width
    cells = []
    for header in headers:
        cell = WriteOnlyCell(sheet, value=_safe_text(header))
        cell.font = HEADER_FONT
        cell.alignment = CELL_ALIGNMENT
        cell.border = HEADER_BORDER
        cells.append(cell)
    sheet.append(cells)
    return sheet


def _data_cell(sheet, value):
    cell = WriteOnlyCell(sheet, value=_safe_text(value))
    cell.font = DATA_FONT
    cell.alignment = CELL_ALIGNMENT
    cell.border = THIN_BORDER
    return cell


def _safe_text(value):
    text = " ".join(str(value if value is not None else "").split())
    return "'{}".format(text) if text.startswith(("=", "+", "-", "@")) else text


def _column_widths(headers: Sequence[str], sample_rows: Sequence[tuple]):
    return [
        min(max(max(_display_width(header), *(_display_width(row[index]) for row in sample_rows if index < len(row))) + 2, 9), 80)
        for index, header in enumerate(headers)
    ]


def _display_width(value):
    return sum(2 if ord(character) > 0xFF else 1 for character in str(value))
