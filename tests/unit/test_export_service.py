from pathlib import Path
from decimal import Decimal

import pytest
from openpyxl import load_workbook
from sqlalchemy import Column, Integer, MetaData, String, Table, create_engine, insert

from dataexport_studio.application.export_service import ExportService
from dataexport_studio.domain.errors import ExportError, ValidationError
from dataexport_studio.domain.models import CustomSqlRequest, ExportRequest, FilterCondition, FilterLogic, FilterOperator, SortDirection
from dataexport_studio.infrastructure.database.sqlalchemy_gateway import validate_custom_sql
from dataexport_studio.infrastructure.excel.workbook_writer import write_workbook


def test_exports_filtered_sqlite_table_with_expected_formatting(tmp_path: Path):
    engine = create_engine("sqlite+pysqlite:///:memory:")
    metadata = MetaData()
    people = Table("people", metadata, Column("id", Integer, primary_key=True), Column("name", String), Column("score", Integer))
    metadata.create_all(engine)
    with engine.begin() as connection:
        connection.execute(insert(people), [{"id": 1, "name": "Ada", "score": 95}, {"id": 2, "name": "Bob", "score": 70}])

    destination = tmp_path / "people.xlsx"
    request = ExportRequest(
        schema=None,
        table="people",
        destination=destination,
        filter_condition=FilterCondition("score", FilterOperator.GREATER_OR_EQUAL, "90"),
    )
    totals = []
    progress = []
    count, _ = ExportService().export(
        engine,
        request,
        on_total=lambda total, planned: totals.append((total, planned)),
        on_progress=progress.append,
    )

    workbook = load_workbook(destination)
    sheet = workbook.active
    assert count == 1
    assert totals == [(1, 1)]
    assert progress[-1] == 1
    assert [cell.value for cell in sheet[1]] == ["id", "name", "score"]
    assert [cell.value for cell in sheet[2]] == [1, "Ada", 95]
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].font.name == "SimSun"
    assert sheet["A1"].fill.fill_type is None
    assert sheet["A1"].border.top.style == "medium"
    assert sheet["A2"].alignment.horizontal == "center"
    assert sheet["A2"].border.left.style == "thin"
    assert sheet["A2"].border.right.color is None
    assert sheet.sheet_view.showGridLines is True
    assert sheet.auto_filter.ref is None
    assert sheet.freeze_panes is None
    assert sheet.column_dimensions["B"].width >= 9


def test_rejects_empty_query_without_creating_a_workbook(tmp_path: Path):
    engine = create_engine("sqlite+pysqlite:///:memory:")
    metadata = MetaData()
    Table("people", metadata, Column("id", Integer, primary_key=True))
    metadata.create_all(engine)
    destination = tmp_path / "empty.xlsx"
    totals = []
    request = ExportRequest(schema=None, table="people", destination=destination)

    with pytest.raises(ExportError, match="查询结果为空"):
        ExportService().export(engine, request, on_total=lambda total, planned: totals.append((total, planned)))

    assert totals == [(0, 0)]
    assert not destination.exists()
    assert not destination.with_suffix(".partial.xlsx").exists()


def test_exports_multiple_conditions_and_sorts_rows_with_serialized_select_values(tmp_path: Path):
    engine = create_engine("sqlite+pysqlite:///:memory:")
    metadata = MetaData()
    people = Table("people", metadata, Column("id", Integer, primary_key=True), Column("name", String), Column("score", Integer))
    metadata.create_all(engine)
    with engine.begin() as connection:
        connection.execute(insert(people), [
            {"id": 1, "name": "Ada", "score": 95},
            {"id": 2, "name": "Annie", "score": 80},
            {"id": 3, "name": "Bob", "score": 90},
        ])

    request = ExportRequest(
        schema=None,
        table="people",
        destination=tmp_path / "multiple.xlsx",
        filter_conditions=(
            FilterCondition("score", FilterOperator.GREATER_THAN, "75"),
            FilterCondition("name", FilterOperator.CONTAINS, "A"),
        ),
        # QComboBox.currentData() returns the string value of str-based enums.
        filter_logic="AND",
        sort_column="score",
        sort_direction="desc",
    )
    ExportService().export(engine, request)

    sheet = load_workbook(request.destination).active
    assert [sheet.cell(row, 2).value for row in range(2, 4)] == ["Ada", "Annie"]


def test_exports_or_connected_conditions(tmp_path: Path):
    engine = create_engine("sqlite+pysqlite:///:memory:")
    metadata = MetaData()
    people = Table("people", metadata, Column("id", Integer, primary_key=True), Column("name", String), Column("score", Integer))
    metadata.create_all(engine)
    with engine.begin() as connection:
        connection.execute(insert(people), [
            {"id": 1, "name": "Ada", "score": 95},
            {"id": 2, "name": "Annie", "score": 80},
            {"id": 3, "name": "Bob", "score": 70},
        ])

    request = ExportRequest(
        schema=None,
        table="people",
        destination=tmp_path / "or.xlsx",
        filter_conditions=(
            FilterCondition("score", FilterOperator.GREATER_THAN, "90"),
            FilterCondition("name", FilterOperator.EQUAL, "Bob"),
        ),
        filter_logic=FilterLogic.OR,
        sort_column="id",
    )
    ExportService().export(engine, request)

    sheet = load_workbook(request.destination).active
    assert [sheet.cell(row, 2).value for row in range(2, 4)] == ["Ada", "Bob"]


def test_exports_formula_like_text_as_plain_text(tmp_path: Path):
    engine = create_engine("sqlite+pysqlite:///:memory:")
    metadata = MetaData()
    audit = Table("audit", metadata, Column("id", Integer, primary_key=True), Column("detail", String))
    metadata.create_all(engine)
    with engine.begin() as connection:
        connection.execute(insert(audit), [{"id": 1, "detail": "=HYPERLINK(\"https://example.com\")"}])

    request = ExportRequest(schema=None, table="audit", destination=tmp_path / "audit.xlsx")
    ExportService().export(engine, request)

    cell = load_workbook(request.destination).active["B2"]
    assert cell.data_type != "f"
    assert cell.value == "'=HYPERLINK(\"https://example.com\")"


def test_preserves_native_numbers_but_keeps_numeric_text_as_text(tmp_path: Path):
    destination = tmp_path / "numbers.xlsx"

    write_workbook(
        destination,
        ["整数", "小数", "文本编号"],
        [(3, Decimal("1500.25"), "00123")],
        max_rows=10,
    )

    sheet = load_workbook(destination).active
    assert sheet["A2"].data_type == "n"
    assert sheet["B2"].data_type == "n"
    assert sheet["A2"].value == 3
    assert sheet["B2"].value == 1500.25
    assert sheet["C2"].data_type == "s"
    assert sheet["C2"].value == "00123"


def test_exports_custom_sql_join_and_cte(tmp_path: Path):
    engine = create_engine("sqlite+pysqlite:///:memory:")
    metadata = MetaData()
    departments = Table("departments", metadata, Column("id", Integer, primary_key=True), Column("name", String))
    people = Table("people", metadata, Column("id", Integer, primary_key=True), Column("name", String), Column("department_id", Integer))
    metadata.create_all(engine)
    with engine.begin() as connection:
        connection.execute(insert(departments), [{"id": 1, "name": "Engineering"}])
        connection.execute(insert(people), [{"id": 1, "name": "Ada", "department_id": 1}])

    request = CustomSqlRequest(
        query="""
            WITH active_people AS (
                SELECT id, name, department_id FROM people
            )
            SELECT p.name AS employee, d.name AS department
            FROM active_people p
            LEFT JOIN departments d ON d.id = p.department_id
        """,
        destination=tmp_path / "custom-query.xlsx",
    )

    count, _ = ExportService().export(engine, request)

    sheet = load_workbook(request.destination).active
    assert count == 1
    assert [cell.value for cell in sheet[1]] == ["employee", "department"]
    assert [cell.value for cell in sheet[2]] == ["Ada", "Engineering"]


@pytest.mark.parametrize(
    "query",
    [
        "SELECT * FROM people; DELETE FROM people",
        "INSERT INTO people (id) VALUES (1)",
        "UPDATE people SET name = 'Ada'",
        "SELECT * INTO copied_people FROM people",
        "EXEC refresh_report",
    ],
)
def test_custom_sql_rejects_non_read_only_statements(query):
    with pytest.raises(ValidationError, match="自定义 SQL"):
        validate_custom_sql(query)
